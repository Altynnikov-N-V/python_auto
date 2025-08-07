from openpyxl import load_workbook

workbook = load_workbook("tmp/For_tests.xlsx")
sheet = workbook.active
print(sheet.cell(row=2, column=3).value)
